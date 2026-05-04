"""Data Sources — file-level health / recency dashboard.

Rendered dynamically each time the modal opens so row counts and file dates
reflect current state. Uses `build_tabs()` (not static `TABS`) so the modal
re-scans the filesystem on every render.
"""

from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path

import dash_mantine_components as dmc
from dash_iconify import DashIconify

from config.settings import (
    DATA_COMPLETE, DATA_DIR, DATA_INCREMENTAL, PRIMARY, PROJECT_ROOT,
)
from ..renderers import body, section


# ---------------------------------------------------------------------------
# Dataset catalog — (display name, type, location, loader name)
# Loader names are looked up at render time so we don't hard-fail if a loader
# is unavailable.
# ---------------------------------------------------------------------------

# Incremental folder → (display name, loader function name)
INCREMENTAL_DATASETS = [
    ("Billing",                   "Billing",                 "load_billing"),
    ("ClinicVisits",              "Clinic Visits",           "load_clinic_visits"),
    ("Courses",                   "Courses",                 "load_courses"),
    ("MachineDowntimeFields",     "Machine Downtime Fields", None),
    ("MachineDowntimeGaps",       "Machine Downtime Gaps",   "load_downtime_gaps"),
    ("Plans",                     "Plans",                   "load_plans"),
    ("Procedures",                "Procedures",              "load_procedures"),
    ("Simulations",               "Simulations",             "load_simulations"),
    ("Treatment",                 "Treatment",               "load_treatment"),
    ("TreatmentDetail",           "Treatment Detail",        "load_treatment_detail"),
    ("WeeklyVisits",              "Weekly Visits",           "load_weekly_visits"),
    ("Workflow",                  "Workflow",                "load_workflow"),
]

# Complete file → (display name, loader function name)
COMPLETE_DATASETS = [
    ("2026 CPT Delivery Audit.csv", "CPT Delivery Audit",    "load_cpt_audit"),
    ("Availability.csv",            "Availability (legacy)", "load_availability"),
    ("Daily Volume - Future.csv",   "Daily Volume (Future)", "load_daily_volume_future"),
    ("Daily Volume - Past.csv",     "Daily Volume (Past)",   "load_daily_volume"),
    ("Machine Errors.csv",          "Machine Errors",        "load_machines"),
    ("Machine Statistics.csv",      "Machine Statistics",    "load_machine_statistics"),
    ("OTV Audit.csv",               "OTV Audit",             "load_otvs"),
    ("Physician Schedule.csv",      "Physician Schedule",    "load_physician_schedule"),
    ("ScheduleUpcoming.csv",        "Schedule Upcoming",     "load_schedule_upcoming"),
    ("Tasks.csv",                   "Tasks",                 "load_tasks"),
]


# ---------------------------------------------------------------------------
# Filesystem scanning
# ---------------------------------------------------------------------------

_FILENAME_DATE_RE = re.compile(r"(\d{8})")


def _parse_date_from_filename(name: str) -> _dt.date | None:
    """Extract a YYYYMMDD date from a filename if present."""
    for match in _FILENAME_DATE_RE.finditer(name):
        try:
            return _dt.datetime.strptime(match.group(1), "%Y%m%d").date()
        except ValueError:
            continue
    return None


def _format_size(nbytes: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if nbytes < 1024:
            return f"{nbytes:.1f} {unit}" if unit != "B" else f"{nbytes} B"
        nbytes /= 1024
    return f"{nbytes:.1f} TB"


def _format_date(d: _dt.date | None) -> str:
    return d.strftime("%Y-%m-%d") if d else "—"


def _count_lines(path: Path) -> int:
    """Fast newline count; subtract 1 for the header."""
    try:
        with open(path, "rb") as f:
            n = sum(chunk.count(b"\n") for chunk in iter(lambda: f.read(1 << 20), b""))
        return max(0, n - 1)  # drop header row
    except Exception:
        return 0


def _count_xlsx_rows(path: Path) -> int | None:
    """Row count for an .xlsx file via openpyxl read-only streaming.

    Streams the active sheet with `iter_rows(values_only=True)` and counts —
    `ws.max_row` is unreliable in read-only mode when the file omits the
    dimension header. Subtracts 1 for the header row. Returns None on failure.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            n = sum(1 for _ in ws.iter_rows(values_only=True))
            return max(0, n - 1)
        finally:
            wb.close()
    except Exception:
        return None


def _latest_incremental(folder: Path) -> dict | None:
    """Find the most recent file in an incremental folder + raw row total."""
    if not folder.is_dir():
        return None
    files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".csv"]
    if not files:
        return None
    # Sort by date-in-filename if present, else by mtime
    def _sort_key(p: Path):
        d = _parse_date_from_filename(p.name)
        return (d or _dt.date.min, p.stat().st_mtime)
    latest = max(files, key=_sort_key)
    stat = latest.stat()
    raw_total = sum(_count_lines(p) for p in files)
    return {
        "path": latest,
        "name": latest.name,
        "size": stat.st_size,
        "mtime": _dt.datetime.fromtimestamp(stat.st_mtime).date(),
        "data_date": _parse_date_from_filename(latest.name),
        "file_count": len(files),
        "raw_rows": raw_total,
    }


def _latest_xlsx_group(prefix: str) -> dict | None:
    """Find all xlsx files in DATA_DIR root matching a given filename prefix.

    Used to separately scan the rad-onc Referrals feed
    (`Referrals_Report_RadiantCare_All_*.xlsx`) from the med-onc PRCS feed
    (`Referrals_Report_PRCS_*.xlsx`) — two manual Excel exports that share
    the `Referrals_Report_` family prefix but represent different datasets.
    """
    if not DATA_DIR.is_dir():
        return None
    files = [
        p for p in DATA_DIR.iterdir()
        if p.is_file() and p.name.startswith(prefix) and p.suffix.lower() == ".xlsx"
    ]
    if not files:
        return None
    latest = max(files, key=lambda p: p.stat().st_mtime)
    stat = latest.stat()
    return {
        "path": latest,
        "paths": files,
        "name": latest.name,
        "size": stat.st_size,
        "mtime": _dt.datetime.fromtimestamp(stat.st_mtime).date(),
        "data_date": _parse_date_from_filename(latest.name),
        "file_count": len(files),
    }


def _status_auto(data_date: _dt.date | None, today: _dt.date) -> tuple[str, str]:
    """Health status for an automated (Tue–Sat 2 AM) file.

    Returns (color, label). Green ≤2 days, yellow 3–6, red ≥7.
    """
    if data_date is None:
        return "gray", "unknown"
    age = (today - data_date).days
    if age <= 2:
        return "green", f"{age}d old"
    if age <= 6:
        return "yellow", f"{age}d old"
    return "red", f"{age}d old"


def _status_manual(data_date: _dt.date | None, today: _dt.date) -> tuple[str, str]:
    """Health status for a manual (Referrals) file. Green <7d, yellow <30d, red ≥30d."""
    if data_date is None:
        return "gray", "unknown"
    age = (today - data_date).days
    if age < 7:
        return "green", f"{age}d old"
    if age < 30:
        return "yellow", f"{age}d old"
    return "red", f"{age}d old"


# ---------------------------------------------------------------------------
# Row counts (via in-memory loaders; fast if already cached)
# ---------------------------------------------------------------------------

def _row_count(loader_name: str | None) -> int | None:
    """Call a loader (if named) and return len(df). Returns None on failure."""
    if not loader_name:
        return None
    try:
        from data import loader as _loader  # deferred to avoid import cycles
        fn = getattr(_loader, loader_name, None)
        if fn is None:
            return None
        df = fn()
        return int(len(df)) if df is not None else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Table row builder
# ---------------------------------------------------------------------------

_STATUS_BG = {
    "green":  "#10B981",
    "yellow": "#F59E0B",
    "red":    "#EF4444",
    "gray":   "#9CA3AF",
}


def _status_dot(color: str, label: str):
    """Bold pill with colored fill; used in the data table status column."""
    return dmc.Box(
        style={
            "display": "inline-flex",
            "alignItems": "center",
            "gap": "6px",
            "backgroundColor": _STATUS_BG.get(color, "#9CA3AF"),
            "color": "#FFFFFF",
            "padding": "4px 10px",
            "borderRadius": "999px",
            "fontSize": "11px",
            "fontWeight": 700,
            "lineHeight": 1,
            "letterSpacing": "0.2px",
            "whiteSpace": "nowrap",
        },
        children=[
            DashIconify(icon="tabler:circle-filled", width=10, color="#FFFFFF"),
            label,
        ],
    )


def _build_row(entry: dict, today: _dt.date, is_manual: bool = False) -> dmc.TableTr:
    status_fn = _status_manual if is_manual else _status_auto
    color, label = status_fn(entry["data_date"], today)
    unique = entry.get("unique_rows")
    raw = entry.get("raw_rows")
    unique_txt = f"{unique:,}" if unique is not None else "—"
    raw_txt = f"{raw:,}" if raw is not None else "—"
    if unique is not None and raw is not None and raw > 0:
        dup = raw - unique
        dup_txt = f"{dup:,} ({dup / raw:.0%})" if dup > 0 else "0"
    else:
        dup_txt = "—"
    files_txt = f"{entry['file_count']:,}" if entry.get("file_count") else "1"
    type_color = {"Incremental": "violet", "Complete": "blue",
                  "Manual": "orange"}.get(entry["type"], "gray")
    return dmc.TableTr(children=[
        dmc.TableTd(_status_dot(color, label)),
        dmc.TableTd(dmc.Text(entry["display_name"], size="xs", fw=500)),
        dmc.TableTd(dmc.Badge(entry["type"], size="xs", variant="light",
                              color=type_color)),
        dmc.TableTd(dmc.Text(_format_date(entry["data_date"]), size="xs")),
        dmc.TableTd(dmc.Text(_format_date(entry["mtime"]), size="xs", c="dimmed")),
        dmc.TableTd(dmc.Text(_format_size(entry["size"]), size="xs", c="dimmed")),
        dmc.TableTd(dmc.Text(files_txt, size="xs", c="dimmed", ta="right")),
        dmc.TableTd(dmc.Text(raw_txt, size="xs", c="dimmed", ta="right")),
        dmc.TableTd(dmc.Text(dup_txt, size="xs", c="dimmed", ta="right")),
        dmc.TableTd(dmc.Text(unique_txt, size="xs", ta="right", fw=500)),
    ])


# ---------------------------------------------------------------------------
# Main content builder (called on each modal render)
# ---------------------------------------------------------------------------

def _build_content():
    today = _dt.date.today()
    entries: list[dict] = []

    # Incremental ---------------------------------------------------------
    for folder_name, display_name, loader_name in INCREMENTAL_DATASETS:
        info = _latest_incremental(DATA_INCREMENTAL / folder_name)
        if info is None:
            entries.append({
                "display_name": display_name, "type": "Incremental",
                "data_date": None, "mtime": None, "size": 0,
                "file_count": 0, "raw_rows": 0, "unique_rows": None,
            })
            continue
        entries.append({
            "display_name": display_name,
            "type": "Incremental",
            "data_date": info["data_date"],
            "mtime": info["mtime"],
            "size": info["size"],
            "file_count": info["file_count"],
            "raw_rows": info["raw_rows"],
            "unique_rows": _row_count(loader_name),
        })

    # Complete ------------------------------------------------------------
    for file_name, display_name, loader_name in COMPLETE_DATASETS:
        path = DATA_COMPLETE / file_name
        if not path.is_file():
            entries.append({
                "display_name": display_name, "type": "Complete",
                "data_date": None, "mtime": None, "size": 0,
                "file_count": 0, "raw_rows": 0, "unique_rows": None,
            })
            continue
        stat = path.stat()
        mtime = _dt.datetime.fromtimestamp(stat.st_mtime).date()
        raw_rows = _count_lines(path)
        entries.append({
            "display_name": display_name,
            "type": "Complete",
            # Complete files have no date-suffix; use mtime as the data date.
            "data_date": mtime,
            "mtime": mtime,
            "size": stat.st_size,
            "file_count": 1,
            "raw_rows": raw_rows,
            "unique_rows": _row_count(loader_name),
        })

    # Manual (two separate xlsx families: rad-onc referrals + med-onc PRCS) -
    for prefix, display, loader in (
        ("Referrals_Report_RadiantCare_All_", "Referrals Report",            "load_referrals"),
        ("Referrals_Report_PRCS_",            "Referrals Report (Med-Onc)",  "load_medonc_referrals"),
    ):
        ref = _latest_xlsx_group(prefix)
        if ref is not None:
            entries.append({
                "display_name": display,
                "type": "Manual",
                "data_date": ref["data_date"] or ref["mtime"],
                "mtime": ref["mtime"],
                "size": ref["size"],
                "file_count": ref["file_count"],
                "raw_rows": sum((_count_xlsx_rows(p) or 0) for p in ref["paths"]),
                "unique_rows": _row_count(loader),
            })
        else:
            entries.append({
                "display_name": display, "type": "Manual",
                "data_date": None, "mtime": None, "size": 0,
                "file_count": 0, "raw_rows": None, "unique_rows": None,
            })

    # Summary stats -------------------------------------------------------
    total_unique = sum((e["unique_rows"] or 0) for e in entries)
    total_raw = sum((e["raw_rows"] or 0) for e in entries)
    total_dup = max(0, total_raw - total_unique) if total_raw > 0 else 0
    dup_rate = (total_dup / total_raw) if total_raw > 0 else 0
    total_size = sum(e["size"] for e in entries)
    auto_entries = [e for e in entries if e["type"] != "Manual"]
    green = sum(1 for e in auto_entries
                if _status_auto(e["data_date"], today)[0] == "green")
    yellow = sum(1 for e in auto_entries
                 if _status_auto(e["data_date"], today)[0] == "yellow")
    red = sum(1 for e in auto_entries
              if _status_auto(e["data_date"], today)[0] == "red")

    # Stats row -----------------------------------------------------------
    stats_row = dmc.Grid(
        gutter="md",
        children=[
            dmc.GridCol(_stat_tile(
                "Unique Rows", f"{total_unique:,}",
                f"{total_raw:,} raw · {total_dup:,} deduped ({dup_rate:.0%})",
            ), span=3),
            dmc.GridCol(_stat_tile("Datasets", f"{len(entries)}",
                                   f"{sum(e['file_count'] for e in entries):,} files"),
                        span=3),
            dmc.GridCol(_stat_tile("Total Size", _format_size(total_size),
                                   "all snapshots"), span=3),
            dmc.GridCol(_stat_tile(
                "Automated Health",
                f"{green} / {len(auto_entries)}",
                f"{yellow} stale, {red} old",
            ), span=3),
        ],
    )

    # Table ---------------------------------------------------------------
    header_cells = ["Status", "Dataset", "Type", "Data Date",
                    "File mtime", "Size", "Files", "Raw", "Duped", "Unique"]
    body_rows = [_build_row(e, today, is_manual=(e["type"] == "Manual"))
                 for e in entries]

    table = dmc.Table(
        striped=True, highlightOnHover=True,
        withTableBorder=True, withColumnBorders=True,
        fz="xs", mt="sm",
        children=[
            dmc.TableThead(
                dmc.TableTr(
                    [dmc.TableTh(h) for h in header_cells]
                ),
            ),
            dmc.TableTbody(body_rows),
        ],
    )

    # Legend & notes ------------------------------------------------------
    legend = dmc.Group(
        gap="md", mt="sm",
        children=[
            _status_dot("green", "Automated ≤2d / Manual <7d"),
            _status_dot("yellow", "Automated 3–6d / Manual <30d"),
            _status_dot("red", "Automated ≥7d / Manual ≥30d"),
        ],
    )

    return dmc.Stack(
        gap="md",
        children=[
            dmc.Text(
                "Health and recency of every CSV / XLSX the dashboard ingests. "
                "Refreshes each time you open this page. Automated reports "
                "(Incremental and Complete) drop in from ARIA Report Builder "
                "overnight Tuesday through Saturday by 2 AM. The Referrals "
                "report is a manual Excel export.",
                size="sm", c="dimmed", style={"lineHeight": 1.6},
            ),

            stats_row,

            section(
                "Data source files",
                "tabler:files",
                body(
                    f"Scanned {_dt.datetime.now().strftime('%Y-%m-%d %H:%M')}. "
                    "Data Date is the date-suffix parsed from the filename for "
                    "incremental / manual exports, or the file mtime for "
                    "Complete refresh files. Rows are counted from the in-"
                    "process loader cache — empty if a loader hasn't been "
                    "touched yet this session.",
                ),
                table,
                legend,
            ),

            section(
                "Expected delivery schedule",
                "tabler:calendar-clock",
                body(
                    "ARIA Report Builder runs Tue–Sat nightly, landing each "
                    "day's prior-weekday data in OneDrive by 2 AM. No Sun / "
                    "Mon deliveries. ≥7 days stale means an export is stuck. "
                    "The Referrals Report is a manual Excel pull — weekly "
                    "is fine, >30 days is stale.",
                ),
            ),

            _build_persisted_section(),
        ],
    )


def _stat_tile(label: str, value: str, sublabel: str | None = None) -> dmc.Paper:
    children = [
        dmc.Text(label, size="xs", c="dimmed", fw=500, tt="uppercase",
                 style={"letterSpacing": "0.5px"}),
        dmc.Text(value, size="lg", c=PRIMARY, fw=700, mt=4),
    ]
    if sublabel:
        children.append(dmc.Text(sublabel, size="xs", c="dimmed", mt=2))
    return dmc.Paper(
        p="md", radius="md", withBorder=True,
        style={"height": "100%"},
        children=children,
    )


# ---------------------------------------------------------------------------
# Persisted app data — SQLite + lookup files + geocode caches
# ---------------------------------------------------------------------------

SQLITE_PATH = PROJECT_ROOT / "reviews.db"
DATA_LOCAL = PROJECT_ROOT / "data"

# Consolidated CMS lookup CSVs the app actually reads at runtime. The raw
# extracted/ yearly files under data/{rvu,opps}_files/extracted/ are only used
# offline to regenerate these lookups, so we report on the lookups instead.
CMS_LOOKUP_FILES = [
    (DATA_LOCAL / "rvu_files" / "rvu_lookup.csv",
     "Physician Fee Schedule (PPRRVU, all years)"),
    (DATA_LOCAL / "rvu_files" / "gpci_rest_of_wa.csv",
     "GPCI — Rest of WA locality"),
    (DATA_LOCAL / "opps_files" / "opps_lookup.csv",
     "OPPS Addendum B APC rates (all years)"),
    (DATA_LOCAL / "opps_files" / "opps_params.csv",
     "OPPS conversion factors / parameters"),
]


# Tables tied to features not currently used — hidden from the Data page.
_HIDDEN_REVIEWS_TABLES = {
    "insurance_rates",        # per-payor rates feature deprecated
    "insurance_rate_history", # audit trail for the above
}


def _scan_reviews_db() -> dict:
    """Return a dialect-aware summary of the reviews database.

    In production the backend is Postgres (Supabase); locally it is SQLite.
    Shape:
        {
          "backend":    "SQLite" | "Postgres",
          "identifier": "reviews.db" | "<schema> schema",
          "sublabel":   short descriptor for the DB File tile,
          "rows":       [{name, rows, reviewed, latest}, ...],
          "size_bytes": int | None,   # SQLite file size; None for Postgres
          "available":  bool,
        }
    """
    from data.reviews_db import _USE_POSTGRES, _PG_SCHEMA, _connect

    if _USE_POSTGRES:
        try:
            with _connect() as conn:
                tables = [r["table_name"] for r in conn.execute(
                    "SELECT table_name FROM information_schema.tables "
                    "WHERE table_schema = ? AND table_type = 'BASE TABLE' "
                    "ORDER BY table_name",
                    (_PG_SCHEMA,),
                ).fetchall()]
                tables = [t for t in tables if t not in _HIDDEN_REVIEWS_TABLES]
                rows: list[dict] = []
                for t in tables:
                    n = conn.execute(f'SELECT COUNT(*) AS c FROM "{t}"').fetchone()["c"]
                    col_rows = conn.execute(
                        "SELECT column_name FROM information_schema.columns "
                        "WHERE table_schema = ? AND table_name = ?",
                        (_PG_SCHEMA, t),
                    ).fetchall()
                    cols = {r["column_name"] for r in col_rows}
                    reviewed_txt = "—"
                    if "reviewed" in cols and n > 0:
                        rc = conn.execute(
                            f'SELECT COUNT(*) AS c FROM "{t}" WHERE reviewed=1'
                        ).fetchone()["c"]
                        pct = (rc / n) if n else 0
                        reviewed_txt = f"{rc:,} / {n:,} ({pct:.0%})"
                    latest = None
                    if "updated_at" in cols and n > 0:
                        m = conn.execute(
                            f'SELECT MAX(updated_at) AS m FROM "{t}"'
                        ).fetchone()["m"]
                        if m:
                            latest = str(m)[:10]
                    rows.append({
                        "name": t, "rows": n,
                        "reviewed": reviewed_txt, "latest": latest or "—",
                    })
            return {
                "backend": "Postgres",
                "identifier": f"{_PG_SCHEMA} schema",
                "sublabel": "Postgres (Supabase)",
                "rows": rows,
                "size_bytes": None,
                "available": True,
            }
        except Exception:
            return {
                "backend": "Postgres",
                "identifier": f"{_PG_SCHEMA} schema",
                "sublabel": "unreachable",
                "rows": [],
                "size_bytes": None,
                "available": False,
            }

    # SQLite (local dev)
    import sqlite3

    if not SQLITE_PATH.is_file():
        return {
            "backend": "SQLite",
            "identifier": SQLITE_PATH.name,
            "sublabel": "not found",
            "rows": [],
            "size_bytes": 0,
            "available": False,
        }
    rows: list[dict] = []
    try:
        con = sqlite3.connect(SQLITE_PATH)
        cur = con.cursor()
        tables = [r[0] for r in cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name NOT LIKE 'sqlite_%' ORDER BY name"
        )]
        tables = [t for t in tables if t not in _HIDDEN_REVIEWS_TABLES]
        for t in tables:
            n = cur.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
            cols = [c[1] for c in cur.execute(f'PRAGMA table_info("{t}")').fetchall()]
            reviewed_txt = "—"
            if "reviewed" in cols and n > 0:
                r = cur.execute(f'SELECT COUNT(*) FROM "{t}" WHERE reviewed=1').fetchone()[0]
                pct = (r / n) if n else 0
                reviewed_txt = f"{r:,} / {n:,} ({pct:.0%})"
            latest = None
            if "updated_at" in cols and n > 0:
                latest = cur.execute(f'SELECT MAX(updated_at) FROM "{t}"').fetchone()[0]
                if latest:
                    latest = latest[:10]
            rows.append({
                "name": t, "rows": n, "reviewed": reviewed_txt, "latest": latest or "—",
            })
        con.close()
    except Exception:
        pass
    stat = SQLITE_PATH.stat()
    mtime = _dt.datetime.fromtimestamp(stat.st_mtime).date()
    return {
        "backend": "SQLite",
        "identifier": SQLITE_PATH.name,
        "sublabel": f"updated {mtime}",
        "rows": rows,
        "size_bytes": stat.st_size,
        "available": True,
    }


def _csv_row_count(path: Path) -> int | None:
    if not path.is_file():
        return None
    return _count_lines(path)


def _build_persisted_section():
    # --- Reviews database (SQLite locally, Postgres in production) -------
    db = _scan_reviews_db()
    db_rows = db["rows"]
    total_db_rows = sum(r["rows"] for r in db_rows)

    db_table = dmc.Table(
        striped=True, highlightOnHover=True,
        withTableBorder=True, withColumnBorders=True,
        fz="xs", mt="sm",
        children=[
            dmc.TableThead(dmc.TableTr([
                dmc.TableTh("Table"),
                dmc.TableTh("Rows"),
                dmc.TableTh("Reviewed"),
                dmc.TableTh("Last Update"),
            ])),
            dmc.TableTbody([
                dmc.TableTr([
                    dmc.TableTd(dmc.Text(r["name"], size="xs", fw=500)),
                    dmc.TableTd(dmc.Text(f"{r['rows']:,}", size="xs", ta="right")),
                    dmc.TableTd(dmc.Text(r["reviewed"], size="xs", c="dimmed")),
                    dmc.TableTd(dmc.Text(r["latest"], size="xs", c="dimmed")),
                ])
                for r in db_rows
            ]),
        ],
    )

    # --- CMS consolidated lookups (what the app actually reads) ----------
    cms_rows = []
    for path, desc in CMS_LOOKUP_FILES:
        if not path.is_file():
            cms_rows.append({
                "path": path, "desc": desc,
                "rows": None, "size": 0, "mtime": None, "present": False,
            })
            continue
        stat = path.stat()
        cms_rows.append({
            "path": path, "desc": desc,
            "rows": _csv_row_count(path),
            "size": stat.st_size,
            "mtime": _dt.datetime.fromtimestamp(stat.st_mtime).date(),
            "present": True,
        })
    cms_present = [r for r in cms_rows if r["present"]]
    cms_total_rows = sum((r["rows"] or 0) for r in cms_present)
    cms_total_bytes = sum(r["size"] for r in cms_present)

    cms_table = dmc.Table(
        striped=True, highlightOnHover=True,
        withTableBorder=True, withColumnBorders=True,
        fz="xs", mt="sm",
        children=[
            dmc.TableThead(dmc.TableTr([
                dmc.TableTh("File"),
                dmc.TableTh("Purpose"),
                dmc.TableTh("Rows"),
                dmc.TableTh("Size"),
                dmc.TableTh("Last Update"),
            ])),
            dmc.TableTbody([
                dmc.TableTr([
                    dmc.TableTd(dmc.Text(
                        str(r["path"].relative_to(PROJECT_ROOT)),
                        size="xs", fw=500,
                    )),
                    dmc.TableTd(dmc.Text(r["desc"], size="xs", c="dimmed")),
                    dmc.TableTd(dmc.Text(
                        f"{r['rows']:,}" if r["rows"] is not None else "—",
                        size="xs", ta="right",
                    )),
                    dmc.TableTd(dmc.Text(
                        _format_size(r["size"]) if r["present"] else "—",
                        size="xs", c="dimmed",
                    )),
                    dmc.TableTd(dmc.Text(
                        _format_date(r["mtime"]) if r["present"] else "missing",
                        size="xs",
                        c="red" if not r["present"] else "dimmed",
                    )),
                ])
                for r in cms_rows
            ]),
        ],
    )

    # --- Lookup CSVs -----------------------------------------------------
    lookup_files = [
        ("diagnosis_subcategories.csv", "Diagnosis subcategory seed"),
        ("zcta_centroids.csv",           "US ZIP centroids (for Patients map)"),
        ("geocode_cache.csv",            "ZIP geocode cache"),
        ("geocode_addr_cache.csv",       "Address geocode cache"),
    ]

    lookup_rows = []
    for fname, desc in lookup_files:
        p = DATA_LOCAL / fname
        if not p.is_file():
            continue
        stat = p.stat()
        mtime = _dt.datetime.fromtimestamp(stat.st_mtime).date()
        rc = _csv_row_count(p)
        lookup_rows.append({
            "name": fname, "desc": desc, "rows": rc,
            "size": stat.st_size, "mtime": mtime,
        })

    lookup_table = dmc.Table(
        striped=True, highlightOnHover=True,
        withTableBorder=True, withColumnBorders=True,
        fz="xs", mt="sm",
        children=[
            dmc.TableThead(dmc.TableTr([
                dmc.TableTh("File"),
                dmc.TableTh("Purpose"),
                dmc.TableTh("Rows"),
                dmc.TableTh("Size"),
                dmc.TableTh("Last Update"),
            ])),
            dmc.TableTbody([
                dmc.TableTr([
                    dmc.TableTd(dmc.Text(r["name"], size="xs", fw=500)),
                    dmc.TableTd(dmc.Text(r["desc"], size="xs", c="dimmed")),
                    dmc.TableTd(dmc.Text(f"{r['rows']:,}" if r["rows"] is not None else "—",
                                         size="xs", ta="right")),
                    dmc.TableTd(dmc.Text(_format_size(r["size"]), size="xs", c="dimmed")),
                    dmc.TableTd(dmc.Text(_format_date(r["mtime"]), size="xs", c="dimmed")),
                ])
                for r in lookup_rows
            ]),
        ],
    )

    # Summary stats row for persisted data -------------------------------
    db_sublabel_parts = [f"{len(db_rows)} tables"]
    if db["size_bytes"] is not None:
        db_sublabel_parts.append(_format_size(db["size_bytes"]))
    db_sublabel = " · ".join(db_sublabel_parts)

    summary = dmc.Grid(
        gutter="md", mb="sm",
        children=[
            dmc.GridCol(_stat_tile(
                f"{db['backend']} Rows", f"{total_db_rows:,}",
                db_sublabel,
            ), span=3),
            dmc.GridCol(_stat_tile(
                "CMS Lookup CSVs", f"{len(cms_present)} / {len(cms_rows)}",
                f"{cms_total_rows:,} rows · {_format_size(cms_total_bytes)}",
            ), span=3),
            dmc.GridCol(_stat_tile(
                "Reference CSVs", f"{len(lookup_rows)}",
                f"{sum((r['rows'] or 0) for r in lookup_rows):,} rows total",
            ), span=3),
            dmc.GridCol(_stat_tile(
                "Reviews DB", db["identifier"], db["sublabel"],
            ), span=3),
        ],
    )

    return section(
        "Persisted app data",
        "tabler:database-heart",
        body(
            "Editable state and reference data that live outside the ARIA "
            "warehouse — reviews are stored in Postgres (Supabase) in "
            "production and SQLite locally; lookup CSVs live in the repo.",
        ),
        summary,

        dmc.Text(f"{db['backend']} — {db['identifier']}",
                 fw=600, size="xs", mt="md", mb=2),
        dmc.Text(
            "Written by the Payor Manager, Referring Physician Manager, "
            "Diagnosis Classification Manager, and the CPT / OTV audit UIs. "
            "Reviewed column shows how much of each table has been "
            "human-validated.",
            size="xs", c="dimmed", mb=2,
        ),
        db_table,

        dmc.Text("CMS fee-schedule lookups", fw=600, size="xs", mt="md", mb=2),
        dmc.Text(
            "Consolidated, all-years CSVs read at runtime by the Billing and "
            "OTV Audit pages. Raw yearly files under data/{rvu,opps}_files/"
            "extracted/ are only used offline to regenerate these lookups — "
            "the app never touches them directly.",
            size="xs", c="dimmed", mb=2,
        ),
        cms_table,

        dmc.Text("Reference CSVs", fw=600, size="xs", mt="md", mb=2),
        dmc.Text(
            "Static or slowly-changing lookup files under data/. Geocode "
            "caches grow on-demand as new patient / referrer addresses are "
            "resolved.",
            size="xs", c="dimmed", mb=2,
        ),
        lookup_table,
    )


# ---------------------------------------------------------------------------
# Entry points consumed by modal.py
# ---------------------------------------------------------------------------

def build_tabs() -> list[dict]:
    """Dynamic tabs spec — re-runs the filesystem scan on every call."""
    return [
        {
            "value": "data",
            "label": "Data Sources",
            "icon": "tabler:database-search",
            "content": _build_content(),
        },
    ]
